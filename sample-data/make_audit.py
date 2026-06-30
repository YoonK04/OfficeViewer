import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

base = os.path.dirname(os.path.abspath(__file__))
wb = Workbook()
ws = wb.active
ws.title = "감사"

# 격자선 끄기
ws.sheet_view.showGridLines = False

thin = Side(style="thin", color="000000")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font = Font(bold=True, color="FFFFFF")

# 회전 텍스트 헤더
labels = ["항목", "1월", "2월", "3월"]
for i, t in enumerate(labels, start=1):
    c = ws.cell(row=1, column=i, value=t)
    c.fill = hdr_fill
    c.font = hdr_font
    c.border = box
    if i == 1:
        c.alignment = Alignment(horizontal="center", vertical="center")
    elif i == 2:
        c.alignment = Alignment(textRotation=45, horizontal="center")   # 45도
    elif i == 3:
        c.alignment = Alignment(textRotation=90, horizontal="center")   # 90도
    else:
        c.alignment = Alignment(textRotation=255, horizontal="center")  # 세로쓰기

# 데이터
data = [["매출", 100, 200, 300], ["원가", 60, 110, 150], ["이익", 40, 90, 150]]
for r, row in enumerate(data, start=2):
    for cidx, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=cidx, value=v)
        cell.border = box
        if cidx >= 2:
            cell.number_format = "#,##0"

# 숨김 열(B = 1월)과 숨김 행(3 = 원가)
ws.column_dimensions["B"].hidden = True
ws.row_dimensions[3].hidden = True

ws.column_dimensions["A"].width = 14
for col in ["B", "C", "D"]:
    ws.column_dimensions[col].width = 8
ws.row_dimensions[1].height = 50

ws["A6"] = "※ B열(1월)과 3행(원가)은 숨김 처리됨"
ws["A6"].font = Font(italic=True, color="C00000")

wb.save(os.path.join(base, "감사테스트.xlsx"))
print("audit sample created")
