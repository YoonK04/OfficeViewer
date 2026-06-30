import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

base = os.path.dirname(os.path.abspath(__file__))
wb = Workbook()
ws = wb.active
ws.title = "Cathode Flow"

thin = Side(style="thin", color="000000")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 사용자 지정 열 너비 (원본처럼 좁고 다양하게)
widths = {"A": 12, "B": 12, "C": 3, "D": 14, "E": 3, "F": 3, "G": 16, "H": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# 제목
ws["A1"] = "① Cathode Plate Process Flow"
ws["A1"].font = Font(bold=True, size=14)

yellow = PatternFill("solid", fgColor="FFFF00")
ws["A3"] = "Receiving Barcode"; ws["A3"].fill = yellow; ws["A3"].font = Font(bold=True)
ws["G3"] = "Product Barcode"; ws["G3"].fill = yellow; ws["G3"].font = Font(bold=True)

def make_box(cell_range, text, top_left):
    ws.merge_cells(cell_range)
    c = ws[top_left]
    c.value = text
    c.alignment = center
    # 병합 범위 전체에 테두리
    for row in ws[cell_range]:
        for cell in row:
            cell.border = box

# 박스들 (병합 + 테두리)
make_box("A4:B5", "Plate Input", "A4")
make_box("D4:D5", "Crush &\nVibration Sort\n(shred)", "D4")
make_box("G4:G5", "Carbon 1\n(final goods)", "G4")
make_box("G7:G8", "Carbon 2\n(final goods)", "G7")
make_box("A10:B11", "Copper foil", "A10")
make_box("A13:B14", "Ton-bag", "A13")
make_box("A16:B17", "Vinyl", "A16")
make_box("A19:B20", "Tape", "A19")

ws["C5"] = "→"
ws["H4"] = "Sale"
ws["H7"] = "Sale"
ws["C11"] = "Sale"
ws["C14"] = "Waste"

# 행 높이 일부 지정
for r in [4, 5, 7, 8, 10, 11]:
    ws.row_dimensions[r].height = 22

wb.save(os.path.join(base, "공정도.xlsx"))
print("flowchart sample created")
