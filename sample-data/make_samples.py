import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

base = os.path.dirname(os.path.abspath(__file__))
sub = os.path.join(base, "2026-매출")
os.makedirs(sub, exist_ok=True)

def make_sales(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "월별매출"

    title_font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="0E639C")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "2026년 상반기 매출 요약"
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["월", "제품A", "제품B", "제품C", "합계"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    data = [
        ["1월", 1200000, 980000, 450000],
        ["2월", 1350000, 1010000, 520000],
        ["3월", 1500000, 1100000, 610000],
        ["4월", 1420000, 1250000, 700000],
        ["5월", 1680000, 1330000, 760000],
        ["6월", 1810000, 1400000, 820000],
    ]
    for r, row in enumerate(data, start=3):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = border
            if ci >= 2:
                cell.number_format = "#,##0"
        total = ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})")
        total.number_format = "#,##0"
        total.font = Font(bold=True)
        total.border = border

    ws.column_dimensions["A"].width = 10
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 14

    ws2 = wb.create_sheet("메모")
    ws2["A1"] = "이 시트는 두 번째 시트입니다."
    ws2["A1"].font = Font(italic=True, color="C00000")

    wb.save(path)

make_sales(os.path.join(base, "매출요약.xlsx"))
make_sales(os.path.join(sub, "지점별매출.xlsx"))

# 간단한 두 번째 파일
wb = Workbook()
ws = wb.active
ws.title = "재고"
for i, h in enumerate(["품목", "수량", "단가"], start=1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="FFE699")
items = [["볼트", 1200, 50], ["너트", 3400, 30], ["와셔", 8000, 10]]
for r, row in enumerate(items, start=2):
    for ci, v in enumerate(row, start=1):
        ws.cell(row=r, column=ci, value=v)
wb.save(os.path.join(base, "재고현황.xlsx"))

print("samples created at", base)
print(os.listdir(base))
print(os.listdir(sub))
