import os, shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

base = os.path.dirname(os.path.abspath(__file__))
wb = Workbook()
ws = wb.active
ws.title = "대용량"

thin = Side(style="thin", color="CCCCCC")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font = Font(bold=True, color="FFFFFF")
alt_fill = PatternFill("solid", fgColor="EEF3FB")

ROWS, COLS = 3000, 12
for c in range(1, COLS + 1):
    cell = ws.cell(row=1, column=c, value=f"열{c}")
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.border = box
    cell.alignment = Alignment(horizontal="center")

for r in range(2, ROWS + 2):
    for c in range(1, COLS + 1):
        cell = ws.cell(row=r, column=c, value=(r * c) % 100000)
        cell.border = box
        cell.number_format = "#,##0"
        if r % 2 == 0:
            cell.fill = alt_fill

path_xlsx = os.path.join(base, "대용량.xlsx")
wb.save(path_xlsx)
# 같은 내용을 .xlsm로도 복제(파싱 경로 비교용)
shutil.copyfile(path_xlsx, os.path.join(base, "대용량.xlsm"))
print(f"heavy created: {ROWS}x{COLS} cells")
